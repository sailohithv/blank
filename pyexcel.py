from openpyxl import Workbook
from openpyxl.styles import Alignment

"""book = Workbook()
sheet = book.active
sheet.merge_cells('A1:B2')
cell = sheet.cell(row=1, column=1)
cell.value="TABLE NAME"
cell.alignment = Alignment(horizontal='center', vertical='center')
book.save("C:\Users\sandya.nandakumar\Desktop\sassignment2\sample.xlsx")"""

book1 = Workbook()
sheet = book1.active

sheet.freeze_panes = 'B2'

book1.save("C:\Users\LOHITH\Desktop\New folder\sample1.xlsx")
